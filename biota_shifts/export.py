"""Excel и PDF отчёты."""
import os
from io import BytesIO
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from biota_shifts.config import APP_DIR
from biota_shifts.constants import (
    HOURS_GRID_NO_PUNCH,
    HOURS_GRID_SUFFIX_OUTSIDE_GRAPH,
    MONTH_NAMES_RU,
    SCHEDULE_CODE_COLORS,
)

def build_pretty_excel(df: pd.DataFrame, sheet_name: str = "Смены") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Сетка и центрирование для всех ячеек
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                cell.border = border

        for idx, col_name in enumerate(df.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].fillna("")])
            cap = 60 if str(col_name) == "Сотрудник" else 40
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, cap)

    output.seek(0)
    return output.getvalue()


def build_schedule_excel(df: pd.DataFrame, sheet_name: str = "График", year: int | None = None, month: int | None = None) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        center = Alignment(horizontal="center", vertical="center")
        weekend_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Подсветка выходных (сб/вс) в строке заголовка
        if year is not None and month is not None:
            # weekday: Пн=0 ... Вс=6
            for col_idx, col_name in enumerate(df.columns, start=1):
                if str(col_name).isdigit():
                    day_num = int(col_name)
                    if 1 <= day_num <= 31:
                        wd = date(year, month, day_num).weekday()
                        if wd in (5, 6):
                            cell = ws.cell(row=1, column=col_idx)
                            cell.fill = weekend_header_fill
                            cell.font = Font(color="000000", bold=True)

        ws.freeze_panes = "D2"
        ws.auto_filter.ref = ws.dimensions

        day_cols = [idx for idx, name in enumerate(df.columns, start=1) if str(name).isdigit()]
        for row in range(2, ws.max_row + 1):
            for col_idx in day_cols:
                cell = ws.cell(row=row, column=col_idx)
                value = str(cell.value).strip().lower() if cell.value is not None else ""
                if value in SCHEDULE_CODE_COLORS:
                    cell.fill = PatternFill(
                        start_color=SCHEDULE_CODE_COLORS[value],
                        end_color=SCHEDULE_CODE_COLORS[value],
                        fill_type="solid",
                    )
                cell.alignment = center

        # Сетка и центрирование для всех ячеек (включая «Сотрудник» и «Код»)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                cell.border = border

        # Размеры под ваши требования
        # openpyxl: ширина в “Excel units”, высота в “points”
        # приблизительно: px ~= 0.75 * pt  => pt ~= px / 0.75
        day_width_px = 30
        day_width_units = int(round((day_width_px - 5) / 7))  # эмпирическая формула Excel width
        row_height_px = 30
        row_height_points = row_height_px / 0.75

        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = row_height_points

        max_employee_len = 0
        if "Сотрудник" in df.columns:
            max_employee_len = max(
                [len("Сотрудник")] + [len(str(v)) for v in df["Сотрудник"].fillna("").astype(str)]
            )

        # openpyxl задаёт ширину в “Excel units” (не в пикселях).
        # Для стандартного Excel это примерно 1 unit ~= 7px при шрифте Calibri ~11.
        employee_width_units = int(round(180 / 7))

        for idx, col_name in enumerate(df.columns, start=1):
            if str(col_name).isdigit():
                width = day_width_units
            elif col_name == "Сотрудник":
                width = employee_width_units
            elif col_name == "Код":
                width = 8
            else:
                width = 10
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    output.seek(0)
    return output.getvalue()


def _resolve_logo_path() -> Path | None:
    env_logo = (os.getenv("BIOTA_LOGO_PATH") or "").strip()
    if env_logo:
        p = Path(env_logo)
        if p.is_file():
            return p
    for p in (APP_DIR / "Logo.png", APP_DIR / "logo.png"):
        if p.is_file():
            return p
    return None


def _resolve_pdf_cyrillic_font() -> tuple[str, Path | None]:
    """(имя шрифта для ReportLab, путь к TTF или None)."""
    env_font = (os.getenv("BIOTA_REPORT_FONT_TTF") or "").strip()
    candidates: list[Path] = []
    if env_font:
        candidates.append(Path(env_font))
    windir = os.environ.get("WINDIR", "C:\\Windows")
    candidates.extend(
        [
            Path(windir) / "Fonts" / "arial.ttf",
            Path(windir) / "Fonts" / "arialuni.ttf",
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
        ]
    )
    for p in candidates:
        try:
            if p.is_file():
                return ("ArialUnicode", p)
        except OSError:
            continue
    return ("Helvetica", None)


def build_stats_pdf(stats_df: pd.DataFrame, employee_name: str, month_start: date) -> bytes:
    """PDF-отчет для печати по статистике сотрудника за месяц."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from PIL import Image

    # Шрифт с поддержкой кириллицы (Windows / Linux или BIOTA_REPORT_FONT_TTF)
    font_name, ttf_path = _resolve_pdf_cyrillic_font()
    if ttf_path is not None:
        pdfmetrics.registerFont(TTFont("ArialUnicode", str(ttf_path)))
        font_name = "ArialUnicode"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    logo_path = _resolve_logo_path()
    if logo_path is not None:
        logo_buf = BytesIO()
        img = Image.open(logo_path)
        # Не переводить в grayscale — иначе теряется альфа и пропадает прозрачный фон PNG
        if img.mode == "P":
            img = img.convert("RGBA")
        elif img.mode == "CMYK":
            img = img.convert("RGBA")
        elif img.mode not in ("RGBA", "RGB"):
            img = img.convert("RGBA")
        img.save(logo_buf, format="PNG")
        logo_buf.seek(0)
        story.append(RLImage(logo_buf, width=24 * mm, height=24 * mm, mask="auto"))
        story.append(Spacer(1, 3 * mm))

    title = f"Статистика сотрудника: {employee_name}"
    subtitle = f"Период: {month_start.strftime('%m.%Y')}"
    title_style = styles["Title"]
    title_style.fontName = font_name
    title_style.fontSize = 14
    story.append(Paragraph(title, title_style))
    sub_style = styles["Normal"]
    sub_style.fontName = font_name
    story.append(Paragraph(subtitle, sub_style))
    story.append(Spacer(1, 4 * mm))

    def _stats_pdf_header_label(col: str) -> str:
        if col == "Пришел":
            return "Пришел (первая)"
        if col == "Ушел":
            return "Ушел (последняя)"
        return str(col)

    cols = list(stats_df.columns)
    table_data = [[_stats_pdf_header_label(c) for c in cols]]
    for _, r in stats_df.iterrows():
        row_out: list[str] = []
        for c in cols:
            v = r.get(c)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                row_out.append("")
            else:
                row_out.append(str(v))
        table_data.append(row_out)

    # Ширина контента ~190 мм при полях 10 мм — колонки под книжный A4
    _default_w = [22, 16, 38, 22, 22, 22, 18, 18]
    if len(cols) == len(_default_w):
        col_w_mm = _default_w
    else:
        w = 190.0 / max(len(cols), 1)
        col_w_mm = [w] * len(cols)
    table = Table(table_data, repeatRows=1, colWidths=[w * mm for w in col_w_mm])
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def build_hours_grid_pdf(grid_df: pd.DataFrame, year: int, month: int) -> bytes:
    """PDF сетки «Часы по дням»: только сотрудник и дни (без Порядок/Код), альбом A4."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from PIL import Image

    font_name, ttf_path = _resolve_pdf_cyrillic_font()
    if ttf_path is not None:
        pdfmetrics.registerFont(TTFont("ArialUnicode", str(ttf_path)))
        font_name = "ArialUnicode"

    buffer = BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    logo_path = _resolve_logo_path()
    if logo_path is not None:
        logo_buf = BytesIO()
        img = Image.open(logo_path)
        if img.mode == "P":
            img = img.convert("RGBA")
        elif img.mode == "CMYK":
            img = img.convert("RGBA")
        elif img.mode not in ("RGBA", "RGB"):
            img = img.convert("RGBA")
        img.save(logo_buf, format="PNG")
        logo_buf.seek(0)
        story.append(RLImage(logo_buf, width=22 * mm, height=22 * mm, mask="auto"))
        story.append(Spacer(1, 2 * mm))

    title_style = styles["Title"]
    title_style.fontName = font_name
    title_style.fontSize = 13
    story.append(Paragraph("Часы по дням (Biota, факт)", title_style))
    sub_style = styles["Normal"]
    sub_style.fontName = font_name
    sub_style.fontSize = 10
    story.append(
        Paragraph(
            f"Период: {MONTH_NAMES_RU.get(month, str(month))} {year}",
            sub_style,
        )
    )
    story.append(
        Paragraph(
            f"Ячейки: время — Biota; «{HOURS_GRID_NO_PUNCH}» — смена д/н без пары отметок; пусто — в графике нет смены; "
            f"время{HOURS_GRID_SUFFIX_OUTSIDE_GRAPH} — факт из Biota, в графике не д/н (нужно отметить смену); от/б/п/кп — по графику.",
            sub_style,
        )
    )
    story.append(Spacer(1, 3 * mm))

    _exclude = frozenset({"Порядок", "Код"})
    _digit_cols = [c for c in grid_df.columns if str(c).isdigit()]
    _digit_sorted = sorted(_digit_cols, key=lambda x: int(str(x)))
    _non_digit = [c for c in grid_df.columns if str(c) not in _exclude and c not in _digit_cols]
    cols = _non_digit + _digit_sorted
    table_data: list[list[str]] = [list(cols)]
    for _, r in grid_df.iterrows():
        row_out: list[str] = []
        for c in cols:
            v = r.get(c, "")
            if v is None or (isinstance(v, float) and pd.isna(v)):
                row_out.append("")
            else:
                row_out.append(str(v))
        table_data.append(row_out)

    day_cols = [c for c in cols if str(c).isdigit()]
    n_day = len(day_cols)
    # Ширина полезной области, мм (A4 альбом 297 − поля)
    content_mm = 297.0 - 16.0
    used_lead_mm = 0.0
    for c in cols:
        if str(c).isdigit():
            continue
        cs = str(c)
        if cs == "Порядок":
            used_lead_mm += 9.0
        elif cs == "Код":
            used_lead_mm += 14.0
        elif cs == "Сотрудник":
            used_lead_mm += 46.0
        else:
            used_lead_mm += 16.0
    rest_mm = max(0.0, content_mm - used_lead_mm)
    day_mm = max(4.2, rest_mm / n_day) if n_day else 5.0
    col_widths_pt: list[float] = []
    for c in cols:
        if str(c).isdigit():
            col_widths_pt.append(day_mm * mm)
        else:
            cs = str(c)
            if cs == "Порядок":
                col_widths_pt.append(9 * mm)
            elif cs == "Код":
                col_widths_pt.append(14 * mm)
            elif cs == "Сотрудник":
                col_widths_pt.append(46 * mm)
            else:
                col_widths_pt.append(16 * mm)

    table = Table(table_data, repeatRows=1, colWidths=col_widths_pt, splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 7),
                ("FONTSIZE", (0, 1), (-1, -1), 6),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# Регламент: шкала 08:00–20:00 (720 мин). Экспорт: 144 колонки по 5 мин; шапка — полуторачасовые метки до 20:00.
REG_TL_START_MIN = 8 * 60
REG_TL_END_MIN = 20 * 60
REG_EXPORT_SLOT_MINUTES = 5
REG_EXPORT_N_SLOTS = (REG_TL_END_MIN - REG_TL_START_MIN) // REG_EXPORT_SLOT_MINUTES
REG_EXPORT_SLOTS_PER_30 = 30 // REG_EXPORT_SLOT_MINUTES

DEPT_CLASS_BORDER_HEX = {
    "dept-c1": "7AA2FF",
    "dept-c2": "78D2B4",
    "dept-c3": "F2B66E",
    "dept-c4": "D89AF5",
    "dept-c5": "FF9F9F",
    "dept-c6": "9FD7FF",
    "dept-c7": "B6DF83",
    "dept-c8": "F0C4FF",
}


def _reg_hm_to_minutes(hm: str) -> int:
    parts = str(hm).strip().split(":")
    return int(parts[0]) * 60 + int(parts[1])


def _reg_slot_overlap_export(start_m: int, end_m: int, slot_i: int) -> bool:
    a = REG_TL_START_MIN + slot_i * REG_EXPORT_SLOT_MINUTES
    b = a + REG_EXPORT_SLOT_MINUTES
    return end_m > a and start_m < b


def _reg_slot_paint_export(row: dict, slot_i: int) -> str:
    """«ln» перекрывает «bf»; слоты по REG_EXPORT_SLOT_MINUTES (5 мин в экспорте)."""
    bf_s = _reg_hm_to_minutes(row["breakfast_start"])
    bf_e = _reg_hm_to_minutes(row["breakfast_end"])
    ln_s = _reg_hm_to_minutes(row["lunch_start"])
    ln_e = _reg_hm_to_minutes(row["lunch_end"])
    if _reg_slot_overlap_export(ln_s, ln_e, slot_i):
        return "ln"
    if _reg_slot_overlap_export(bf_s, bf_e, slot_i):
        return "bf"
    return ""


def _reg_shift_caption(shift: str) -> str:
    return "Ночная смена" if shift == "н" else "Дневная смена"


def _pdf_interval_paragraph_xml(start: str, end: str) -> str:
    """В узкой ячейке одна строка «09:00–09:30» ломается; делаем три короткие строки."""
    return (
        f'<font size="6.5">{start}</font><br/>'
        f'<font size="5" color="#5A6578">—</font><br/>'
        f'<font size="6.5">{end}</font>'
    )


def _excel_interval_text(start: str, end: str, ncols: int) -> str:
    """Одна строка, если ширина ≥ ~30 мин по сетке 5 мин (6 колонок)."""
    if ncols >= REG_EXPORT_SLOTS_PER_30:
        return f"{start}–{end}"
    return f"{start}\n–\n{end}"


def build_regulations_timeline_excel(
    rows: list[dict],
    plan_date: date,
    shift: str,
) -> bytes:
    """Светлая вёрстка: 144 колонки по 5 мин (08:00–20:00), шапка — полчаса до 19:30–20:00."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Регламент"

    grid = Side(style="thin", color="888888")
    border = Border(left=grid, right=grid, top=grid, bottom=grid)
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    title_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    hdr_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    bf_fill = PatternFill(start_color="E8EEF9", end_color="E8EEF9", fill_type="solid")
    ln_fill = PatternFill(start_color="E8F4EA", end_color="E8F4EA", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.cell(1, 1, value="Регламент: завтрак и обед · шкала 08:00–20:00 (колонка = 5 мин)")
    ws.cell(1, 1).font = Font(bold=True, size=12, color="000000")
    ws.cell(1, 1).fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=REG_EXPORT_N_SLOTS + 1)
    ws.cell(2, 1, value=f"{plan_date.strftime('%d.%m.%Y')} · {_reg_shift_caption(shift)}")
    ws.cell(2, 1).font = Font(size=10, color="333333")
    ws.cell(2, 1).fill = title_fill
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=REG_EXPORT_N_SLOTS + 1)

    r0 = 4
    ws.cell(r0, 1, value="Сотрудник")
    ws.cell(r0, 1).font = Font(bold=True, size=9, color="000000")
    ws.cell(r0, 1).fill = hdr_fill
    ws.cell(r0, 1).alignment = left_wrap
    ws.cell(r0, 1).border = border
    for g in range(24):
        si = g * REG_EXPORT_SLOTS_PER_30
        c_lo = 2 + si
        c_hi = 2 + si + REG_EXPORT_SLOTS_PER_30 - 1
        ws.merge_cells(start_row=r0, start_column=c_lo, end_row=r0, end_column=c_hi)
        total_min = REG_TL_START_MIN + g * 30
        h, mi = divmod(total_min, 60)
        lbl = f"{h:02d}:{mi:02d}"
        if g == 23:
            lbl = "19:30–20:00"
        cell = ws.cell(r0, c_lo, value=lbl)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, size=9, color="000000")
        cell.alignment = center
        cell.border = border

    data_start = r0 + 1
    for ri, row in enumerate(rows):
        r = data_start + ri
        name_cell = ws.cell(r, 1, value=row["employee_name"])
        name_cell.alignment = left_wrap
        name_cell.font = Font(size=10, color="000000")
        name_cell.fill = white
        dc = row.get("department_class") or "dept-c1"
        hex_b = DEPT_CLASS_BORDER_HEX.get(dc, "7AA2FF")
        name_cell.border = Border(
            left=Side(style="medium", color=hex_b),
            right=grid,
            top=grid,
            bottom=grid,
        )

        covered = [False] * REG_EXPORT_N_SLOTS
        si = 0
        while si < REG_EXPORT_N_SLOTS:
            kind = _reg_slot_paint_export(row, si)
            if not kind:
                si += 1
                continue
            sj = si
            while sj < REG_EXPORT_N_SLOTS and _reg_slot_paint_export(row, sj) == kind:
                sj += 1
            c_lo = 2 + si
            c_hi = 2 + sj - 1
            if c_hi >= c_lo:
                ws.merge_cells(start_row=r, start_column=c_lo, end_row=r, end_column=c_hi)
                top = ws.cell(r, c_lo)
                top.border = border
                ncols = c_hi - c_lo + 1
                if kind == "bf":
                    top.value = _excel_interval_text(
                        row["breakfast_start"], row["breakfast_end"], ncols
                    )
                    top.fill = bf_fill
                    top.font = Font(size=9, color="1A2D50", bold=True)
                else:
                    top.value = _excel_interval_text(
                        row["lunch_start"], row["lunch_end"], ncols
                    )
                    top.fill = ln_fill
                    top.font = Font(size=9, color="143D22", bold=True)
                top.alignment = center
            for i in range(si, sj):
                covered[i] = True
            si = sj

        for si in range(REG_EXPORT_N_SLOTS):
            if covered[si]:
                continue
            c = ws.cell(r, 2 + si, value="")
            c.fill = white
            c.border = border

    ws.column_dimensions["A"].width = 24
    for k in range(REG_EXPORT_N_SLOTS):
        ws.column_dimensions[get_column_letter(2 + k)].width = 0.95

    ws.freeze_panes = ws.cell(data_start, 1)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_regulations_timeline_pdf(
    rows: list[dict],
    plan_date: date,
    shift: str,
) -> bytes:
    """Светлый PDF: 144 колонки по 5 мин до 20:00; шапка — полчаса, последняя метка 19:30–20:00."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, TableStyle, Paragraph, Spacer
    from reportlab.platypus.tables import LongTable

    font_name, ttf_path = _resolve_pdf_cyrillic_font()
    if ttf_path is not None:
        pdfmetrics.registerFont(TTFont("ArialUnicode", str(ttf_path)))
        font_name = "ArialUnicode"

    buffer = BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    title_st = ParagraphStyle(
        "RegTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=14,
        textColor=colors.HexColor("#111111"),
        spaceAfter=4,
    )
    sub_st = ParagraphStyle(
        "RegSub",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9.5,
        textColor=colors.HexColor("#444444"),
        leading=12,
        spaceAfter=4,
    )
    sub_st2 = ParagraphStyle(
        "RegSub2",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        textColor=colors.HexColor("#444444"),
        leading=12,
        spaceAfter=10,
    )
    story.append(Paragraph("Регламент завтрака и обеда", title_st))
    # Два абзаца: длинная строка в одном Paragraph обрезалась у края («20:0»).
    story.append(
        Paragraph(
            f"<b>{plan_date.strftime('%d.%m.%Y')}</b> · {_reg_shift_caption(shift)}",
            sub_st,
        )
    )
    story.append(
        Paragraph(
            "Шкала 08:00–20:00. Каждая колонка таблицы — 5&nbsp;мин; "
            "в шапке — полчаса, последняя полоса до 20:00. "
            "В цветных ячейках — время начала и конца обеда или завтрака.",
            sub_st2,
        )
    )

    name_w = 26 * mm
    content_w = (297 - 16) * mm - name_w
    slot_w = content_w / REG_EXPORT_N_SLOTS
    col_widths = [name_w] + [slot_w] * REG_EXPORT_N_SLOTS

    # Время в шапке — обычные строки (не Paragraph): в узких ячейках Paragraph ломал «08:00» на две строки.
    hdr_st = ParagraphStyle(
        "h0",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        alignment=TA_LEFT,
        leading=10,
        textColor=colors.HexColor("#111111"),
    )
    hdr: list = [Paragraph("<b>Сотрудник</b>", hdr_st)]
    for g in range(24):
        for j in range(REG_EXPORT_SLOTS_PER_30):
            if j == 0:
                total_min = REG_TL_START_MIN + g * 30
                h, mi = divmod(total_min, 60)
                lbl = f"{h:02d}:{mi:02d}"
                if g == 23:
                    lbl = "19:30–20:00"
                hdr.append(lbl)
            else:
                hdr.append("")

    data: list[list] = [hdr]
    name_ps = ParagraphStyle(
        "rn",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        alignment=TA_LEFT,
        leading=10,
        textColor=colors.HexColor("#111111"),
    )
    bf_ps = ParagraphStyle(
        "rbf",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=6.5,
        alignment=TA_CENTER,
        leading=8,
        textColor=colors.HexColor("#0D2847"),
    )
    ln_ps = ParagraphStyle(
        "rln",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=6.5,
        alignment=TA_CENTER,
        leading=8,
        textColor=colors.HexColor("#143D22"),
    )

    ts: list[tuple] = [
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (-1, 0), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111111")),
    ]

    for g in range(24):
        c0 = 1 + g * REG_EXPORT_SLOTS_PER_30
        c1 = c0 + REG_EXPORT_SLOTS_PER_30 - 1
        ts.append(("SPAN", (c0, 0), (c1, 0)))
        ts.append(("BACKGROUND", (c0, 0), (c1, 0), colors.HexColor("#EFEFEF")))

    row_h = 10 * mm
    hdr_h = 7 * mm
    for ri, row in enumerate(rows):
        r = 1 + ri
        cells: list = [Paragraph(row["employee_name"], name_ps)] + [""] * REG_EXPORT_N_SLOTS
        si = 0
        while si < REG_EXPORT_N_SLOTS:
            kind = _reg_slot_paint_export(row, si)
            if not kind:
                si += 1
                continue
            sj = si
            while sj < REG_EXPORT_N_SLOTS and _reg_slot_paint_export(row, sj) == kind:
                sj += 1
            c0 = 1 + si
            c1 = 1 + sj - 1
            ncols = c1 - c0 + 1
            if kind == "bf":
                s, e = row["breakfast_start"], row["breakfast_end"]
                if ncols >= 12:
                    lab = f'<font size="7.5">{s}–{e}</font>'
                else:
                    lab = _pdf_interval_paragraph_xml(s, e)
                cells[1 + si] = Paragraph(lab, bf_ps)
                ts.append(("SPAN", (c0, r), (c1, r)))
                ts.append(("BACKGROUND", (c0, r), (c1, r), colors.HexColor("#E8EEF9")))
            else:
                s, e = row["lunch_start"], row["lunch_end"]
                if ncols >= 12:
                    lab = f'<font size="7.5">{s}–{e}</font>'
                else:
                    lab = _pdf_interval_paragraph_xml(s, e)
                cells[1 + si] = Paragraph(lab, ln_ps)
                ts.append(("SPAN", (c0, r), (c1, r)))
                ts.append(("BACKGROUND", (c0, r), (c1, r), colors.HexColor("#E8F4EA")))
            si = sj
        data.append(cells)
        dc = row.get("department_class") or "dept-c1"
        hx = DEPT_CLASS_BORDER_HEX.get(dc, "7AA2FF")
        ts.append(("LINEBEFORE", (0, r), (0, r), 2.5, colors.HexColor("#" + hx)))

    row_heights = [hdr_h] + [row_h] * len(rows)
    tbl = LongTable(
        data,
        repeatRows=1,
        colWidths=col_widths,
        rowHeights=row_heights,
        splitByRow=1,
    )
    ts.extend(
        [
            ("ROWBACKGROUNDS", (0, 1), (0, -1), [colors.white, colors.HexColor("#F7F7F7")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BBBBBB")),
        ]
    )
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def build_regulations_list_pdf(
    rows: list[dict],
    plan_date: date,
    shift: str,
) -> bytes:
    """PDF регламента в табличном виде: №, сотрудник, завтрак, обед (как в образце)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

    font_name, ttf_path = _resolve_pdf_cyrillic_font()
    if ttf_path is not None:
        pdfmetrics.registerFont(TTFont("ArialUnicode", str(ttf_path)))
        font_name = "ArialUnicode"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        "RegListTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#111111"),
        spaceAfter=2 * mm,
    )
    sub_style = ParagraphStyle(
        "RegListSub",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#222222"),
        spaceAfter=1 * mm,
    )
    info_style = ParagraphStyle(
        "RegListInfo",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#3a3a3a"),
        spaceAfter=4 * mm,
    )

    story.append(Paragraph("Регламент: питания", title_style))
    story.append(
        Paragraph(
            f"{_reg_shift_caption(shift)} · шкала учёта 08:00–20:00",
            sub_style,
        )
    )
    story.append(Paragraph(f"Действует с {plan_date.strftime('%d.%m.%Y')}", sub_style))
    story.append(
        Paragraph(
            "В таблице указаны только временные окна (начало и конец перерыва). "
            "Колонка «Завтрак» — первый интервал приёма пищи, «Обед» — второй.",
            info_style,
        )
    )
    story.append(
        Paragraph(
            "Соблюдайте график, чтобы не пересекаться на кухне и сохранять ритм смены.",
            info_style,
        )
    )
    story.append(Spacer(1, 2 * mm))

    header = [
        "№",
        "Сотрудник",
        "Завтрак (1-й интервал)",
        "Обед (2-й интервал)",
    ]
    table_data: list[list[str]] = [header]
    for i, r in enumerate(rows, start=1):
        b = f"{r['breakfast_start']}–{r['breakfast_end']}"
        l = f"{r['lunch_start']}–{r['lunch_end']}"
        table_data.append([str(i), str(r["employee_name"]), b, l])

    col_widths = [12 * mm, 78 * mm, 42 * mm, 42 * mm]
    table = LongTable(table_data, repeatRows=1, colWidths=col_widths, splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFEFEF")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111111")),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#AAAAAA")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

