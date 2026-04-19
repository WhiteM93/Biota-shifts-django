"""Excel-графики: загрузка, нормализация, пути."""
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Side

from biota_shifts.config import SCHEDULE_DIR
from biota_shifts.constants import SCHEDULE_CODES

# Три последних дня предыдущего месяца (слева от «1» текущего месяца).
PREV_MONTH_KEYS = ("p1", "p2", "p3")


def prev_month_tail_dates(year: int, month: int) -> tuple[date, date, date]:
    first = date(year, month, 1)
    d0 = first - timedelta(days=3)
    d1 = first - timedelta(days=2)
    d2 = first - timedelta(days=1)
    return d0, d1, d2


def is_schedule_day_column(name) -> bool:
    s = str(name).strip()
    if s in PREV_MONTH_KEYS:
        return True
    return s.isdigit() and s.isascii()


def sort_schedule_day_columns(columns: list, year: int, month: int) -> list:
    """Порядок колонок: p1,p2,p3, затем 1..N текущего месяца (только существующие в columns)."""
    colset = {str(c) for c in columns}
    prev = [k for k in PREV_MONTH_KEYS if k in colset]
    digits = sorted([c for c in columns if str(c).isdigit()], key=lambda x: int(str(x)))
    return prev + digits


def schedule_column_to_date(col_key: str, year: int, month: int) -> date | None:
    """Календарная дата для колонки графика (или None)."""
    s = str(col_key).strip()
    if s in PREV_MONTH_KEYS:
        d0, d1, d2 = prev_month_tail_dates(year, month)
        return {PREV_MONTH_KEYS[0]: d0, PREV_MONTH_KEYS[1]: d1, PREV_MONTH_KEYS[2]: d2}[s]
    if s.isdigit():
        di = int(s)
        try:
            return date(year, month, di)
        except ValueError:
            return None
    return None


def available_schedule_years() -> list[int]:
    years = set()
    for p in SCHEDULE_DIR.glob("schedule_*.xlsx"):
        parts = p.stem.split("_")
        if len(parts) >= 3 and parts[1].isdigit():
            years.add(int(parts[1]))
    current_year = datetime.now().year
    years.update({current_year - 1, current_year, current_year + 1, 2026})
    return sorted(years, reverse=True)
def month_bounds(selected_month: date) -> tuple[date, date]:
    start = selected_month.replace(day=1)
    if start.month == 12:
        next_month = date(start.year + 1, 1, 1)
    else:
        next_month = date(start.year, start.month + 1, 1)
    end = next_month - timedelta(days=1)
    return start, end
def schedule_path(year: int, month: int) -> Path:
    return SCHEDULE_DIR / f"schedule_{year}_{month:02d}.xlsx"


def employee_label_row(r: pd.Series) -> str:
    last = str(r.get("last_name", "")).strip()
    first = str(r.get("first_name", "")).strip()
    init = first[:1].upper() if first else ""
    fio = f"{last} {init}." if last and init else (last if last else (init + "." if init else "Без имени"))
    return fio


def _schedule_day_cols(year: int, month: int) -> list[str]:
    days_in_month = (date(year + (month // 12), (month % 12) + 1, 1) - timedelta(days=1)).day
    return list(PREV_MONTH_KEYS) + [str(d) for d in range(1, days_in_month + 1)]


def sanitize_schedule_cell(v) -> str:
    if pd.isna(v):
        return ""
    s = str(v).strip().lower()
    return s if s in SCHEDULE_CODES else ""


def empty_schedule_from_db(employees_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    day_cols = _schedule_day_cols(year, month)
    base = employees_df.copy()
    base["Порядок"] = range(1, len(base) + 1)
    base["Сотрудник"] = base.apply(employee_label_row, axis=1)
    result = base[["Порядок", "emp_code", "Сотрудник"]].rename(columns={"emp_code": "Код"})
    result["Код"] = result["Код"].astype(str)
    for col in day_cols:
        result[col] = ""
    return result.sort_values(["Порядок", "Код"]).reset_index(drop=True)


def normalize_schedule_excel(xl: pd.DataFrame, employees_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    """График только по сотрудникам из БД. Строки с кодами не из БД отбрасываются."""
    day_cols = _schedule_day_cols(year, month)
    label_by_code = {str(r["emp_code"]): employee_label_row(r) for _, r in employees_df.iterrows()}
    valid = set(employees_df["emp_code"].astype(str))

    if "Код" not in xl.columns:
        raise ValueError("В таблице должна быть колонка «Код» (код сотрудника, как в БД).")

    xl = xl.copy()
    xl["Код"] = xl["Код"].astype(str)
    xl = xl[xl["Код"].isin(valid)].copy()
    if "Порядок" not in xl.columns:
        xl["Порядок"] = range(1, len(xl) + 1)
    if "Сотрудник" not in xl.columns:
        xl["Сотрудник"] = xl["Код"]
    for col in day_cols:
        if col not in xl.columns:
            xl[col] = ""
    keep_cols = ["Порядок", "Код", "Сотрудник"] + day_cols
    xl = xl[keep_cols]

    def empty_row_for_code(code: str, order: int) -> dict:
        row = {"Порядок": order, "Код": code, "Сотрудник": label_by_code.get(code, code)}
        for c in day_cols:
            row[c] = ""
        return row

    seen: set[str] = set()
    rows: list[dict] = []
    order = 1
    for _, r in xl.sort_values(["Порядок", "Код"]).iterrows():
        code = r["Код"]
        if code in seen:
            continue
        seen.add(code)
        row = {c: r.get(c, "") for c in keep_cols}
        row["Порядок"] = order
        row["Сотрудник"] = label_by_code.get(code, row.get("Сотрудник", code))
        for c in day_cols:
            row[c] = sanitize_schedule_cell(row.get(c, ""))
        rows.append(row)
        order += 1
    for _, er in employees_df.iterrows():
        c = str(er["emp_code"])
        if c not in seen:
            rows.append(empty_row_for_code(c, order))
            seen.add(c)
            order += 1
    result = pd.DataFrame(rows)
    return result[keep_cols].reset_index(drop=True)


def read_schedule_sheet_from_bytes(data: bytes) -> pd.DataFrame:
    """Читает лист «График» или первый лист."""
    bio = BytesIO(data)
    try:
        return pd.read_excel(bio, sheet_name="График")
    except ValueError:
        bio.seek(0)
        return pd.read_excel(bio, sheet_name=0)


def build_schedule_template_bytes(employees_df: pd.DataFrame, year: int, month: int) -> bytes:
    df = empty_schedule_from_db(employees_df, year, month)
    hint = pd.DataFrame(
        {
            "Код": ["д", "н", "от", "б", "п", "кп", "(пусто)"],
            "Смысл": [
                "дневная смена",
                "ночная смена",
                "отпуск",
                "больничный",
                "прогул",
                "компенсация",
                "смены нет",
            ],
        }
    )
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="График")
        hint.to_excel(writer, index=False, sheet_name="Справка")
        ws = writer.sheets["График"]

        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Под размеры: дни по ~30px, высота строк ~30px
        day_width_px = 30
        day_width_units = int(round((day_width_px - 5) / 7))
        row_height_px = 30
        row_height_points = row_height_px / 0.75

        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = row_height_points

        for ridx, _ in enumerate(df.columns, start=1):
            col_letter = ws.cell(row=1, column=ridx).column_letter
            if is_schedule_day_column(df.columns[ridx - 1]):
                ws.column_dimensions[col_letter].width = day_width_units

        # центр + обводка для ячеек шаблона
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                cell.border = border
    out.seek(0)
    return out.getvalue()


def _read_schedule_dataframe(employees_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    """Чтение и нормализация графика за месяц без подстановки хвоста из прошлого месяца."""
    file_path = schedule_path(year, month)
    if file_path.exists():
        try:
            xl = pd.read_excel(file_path, sheet_name="График")
        except ValueError:
            xl = pd.read_excel(file_path, sheet_name=0)
        return normalize_schedule_excel(xl, employees_df, year, month)
    return empty_schedule_from_db(employees_df, year, month)


def apply_prev_month_tail_from_previous_schedule(
    df: pd.DataFrame, employees_df: pd.DataFrame, year: int, month: int
) -> pd.DataFrame:
    """Заполняет колонки p1–p3 значениями из соответствующих календарных дней предыдущего месяца."""
    out = df.copy()
    d0, d1, d2 = prev_month_tail_dates(year, month)
    yp, mp = d0.year, d0.month
    prev_df = _read_schedule_dataframe(employees_df, yp, mp)
    day_src = [str(d0.day), str(d1.day), str(d2.day)]
    for pk, src_col in zip(PREV_MONTH_KEYS, day_src):
        if pk not in out.columns:
            continue
        if src_col not in prev_df.columns:
            out[pk] = ""
            continue
        cmap: dict[str, str] = {}
        for _, r in prev_df.iterrows():
            code = str(r.get("Код", "")).strip()
            if code:
                cmap[code] = sanitize_schedule_cell(r.get(src_col))
        for idx, row in out.iterrows():
            code = str(row.get("Код", "")).strip()
            out.at[idx, pk] = cmap.get(code, "")
    return out


def load_schedule_table(employees_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    df = _read_schedule_dataframe(employees_df, year, month)
    return apply_prev_month_tail_from_previous_schedule(df, employees_df, year, month)


def save_schedule_table(df: pd.DataFrame, year: int, month: int) -> Path:
    file_path = schedule_path(year, month)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="График")
    return file_path
